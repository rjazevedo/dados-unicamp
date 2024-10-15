import os
import pandas as pd
from pathlib import Path
import yaml
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from unidecode import unidecode


stream = open("capes/configuration.yaml")
config = yaml.safe_load(stream)


def clean_name(name):
    """
    Limpa o nome fornecido, removendo acentos, convertendo para maiúsculas 
    e eliminando espaços em branco extras.

    Parâmetros:
        name (str): O nome a ser limpo.

    Retorna:
        str: O nome limpo, ou uma string vazia se o nome for nulo.
    """
    if pd.isnull(name):
        return ""
    else:
        s = unidecode(name).upper().strip()
        return " ".join(s.split())


def get_all_files(path):
    """Retorna uma lista de todos os arquivos no diretório especificado.

    Args:
        path (str): O caminho do diretório a ser explorado.

    Returns:
        list: Lista de caminhos dos arquivos encontrados no diretório.
    """
    return [f.path for f in os.scandir(path) if f.is_file()]


def list_dirs_capes_input():
    """Lista os diretórios de entrada especificados na configuração.

    Returns:
        list: Lista de caminhos dos diretórios de entrada.
    """
    return [f.path for f in os.scandir(config["path_input"]) if f.is_dir()]


def read_files(year, version):
    """Lê todos os arquivos nos diretórios de entrada e imprime os caminhos dos arquivos encontrados.
    
    Args:
        year (int): O ano dos arquivos.
        version (str): A versão dos arquivos ('2021' ou '2023').
        
    Returns:
        list: Lista de arquivos.
    """
    files = []
    capes_folders = sorted(list_dirs_capes_input())

    for folder in capes_folders:
        all_files = get_all_files(folder)
        for file in all_files:
            if f"br-capes-colsucup-discentes-{year}-{version}.csv" in file:
                files.append(file)
    
    return files


def extract_columns(file):
    df = pd.read_csv(file, nrows=0, encoding='utf-8', low_memory=False)
    return df.columns.tolist()


def compare_columns(columns_2021, columns_2023):
    """Compara as colunas dos arquivos de 2021 e 2023.
    
    Args:
        columns_2021 (list): Lista de nomes das colunas do arquivo de 2021.
        columns_2023 (list): Lista de nomes das colunas do arquivo de 2023.
        
    Returns:
        tuple: Tupla contendo as listas de colunas faltantes, extras e alteradas.
    """
    columns_2021_set = set(columns_2021)
    columns_2023_set = set(columns_2023)
    
    missing_columns = columns_2021_set - columns_2023_set
    extra_columns = columns_2023_set - columns_2021_set
    common_columns = columns_2021_set & columns_2023_set
    
    altered_columns = [col for col in common_columns if columns_2021.index(col) != columns_2023.index(col)]
    
    print(f"Comparison results - Missing: {missing_columns}, Extra: {extra_columns}, Altered: {altered_columns}")
    return list(common_columns), list(missing_columns), list(extra_columns), altered_columns


def print_discrepancies(year, file_2021, file_2023, common_columns, missing_columns, extra_columns, altered_columns):
    """Imprime as discrepâncias entre as colunas dos arquivos de 2021 e 2023.
    
    Args:
        year (int): O ano do arquivo.
        file_2021 (str): O nome do arquivo de 2021.
        file_2023 (str): O nome do arquivo de 2023.
        common_columns (list): A lista de colunas que são comuns entre 2021 e 2023.
        missing_columns (list): A lista de colunas que estão presentes em 2021 mas faltando em 2023.
        extra_columns (list): A lista de colunas que estão presentes em 2023 mas faltando em 2021.
        altered_columns (list): A lista de colunas que foram alteradas.
    """
    rows = []
    max_length = max(len(common_columns), len(missing_columns), len(extra_columns), len(altered_columns))
    
    for i in range(max_length):
        row = {
            "Colunas Iguais": common_columns[i] if i < len(common_columns) else '',
            "Colunas Faltantes": missing_columns[i] if i < len(missing_columns) else '',
            "Colunas Extras": extra_columns[i] if i < len(extra_columns) else '',
            "Colunas Alteradas": altered_columns[i] if i < len(altered_columns) else ''
        }
        rows.append(row)
    
    # Criar o DataFrame a partir das linhas
    df = pd.DataFrame(rows)
    
    print(f"Year: {year}, File 2021: {file_2021}, File 2023: {file_2023}")
    print(df)
    print("\n")
    
    # Diretório para salvar os arquivos
    os.makedirs("/home/giovani/testes", exist_ok=True)
    
    # Salvar o DataFrame em um arquivo Excel
    excel_filename = os.path.join("/home/giovani/testes", f"discrepancies_{year}.xlsx")
    df.to_excel(excel_filename, index=False)
    
    # Ajustar a largura das colunas e a altura das linhas
    wb = load_workbook(excel_filename)
    ws = wb.active
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                max_height = max(max_height, len(str(cell.value).split('\n')))
        ws.row_dimensions[row[0].row].height = max_height * 15
    
    wb.save(excel_filename)
    print(f"Discrepancies saved to {excel_filename}")


def compare_categorical_columns(df_2021, df_2023, columns):
    """Compara as categorias únicas das colunas especificadas entre os DataFrames de 2021 e 2023.
    
    Args:
        df_2021 (DataFrame): DataFrame do arquivo de 2021.
        df_2023 (DataFrame): DataFrame do arquivo de 2023.
        columns (list): Lista de colunas categóricas a serem comparadas.
        
    Returns:
        dict: Dicionário contendo as discrepâncias para cada coluna.
    """
    discrepancies = {}
    for column in columns:
        if column in df_2021.columns and column in df_2023.columns:
            unique_2021 = sorted(set(df_2021[column].unique()))
            unique_2023 = sorted(set(df_2023[column].unique()))
            
            missing_categories = set(unique_2021) - set(unique_2023)
            new_categories = set(unique_2023) - set(unique_2021)
            
            discrepancies[column] = {
                "Categorias Faltantes": list(missing_categories),
                "Novas Categorias": list(new_categories)
            }
    return discrepancies


def print_categorical_discrepancies(year, file_2021, file_2023, discrepancies):
    """Imprime as discrepâncias entre as categorias das colunas dos arquivos de 2021 e 2023.
    
    Args:
        year (int): O ano do arquivo.
        file_2021 (str): O nome do arquivo de 2021.
        file_2023 (str): O nome do arquivo de 2023.
        discrepancies (dict): Dicionário contendo as discrepâncias para cada coluna.
    """
    rows = []
    for column, discrepancy in discrepancies.items():
        max_length = max(len(discrepancy["Categorias Faltantes"]), len(discrepancy["Novas Categorias"]))
        for i in range(max_length):
            row = {
                "Coluna": column,
                "Categorias Faltantes": discrepancy["Categorias Faltantes"][i] if i < len(discrepancy["Categorias Faltantes"]) else '',
                "Novas Categorias": discrepancy["Novas Categorias"][i] if i < len(discrepancy["Novas Categorias"]) else ''
            }
            rows.append(row)
    
    # Criar o DataFrame a partir das linhas
    df = pd.DataFrame(rows)
    
    print(f"Year: {year}, File 2021: {file_2021}, File 2023: {file_2023}")
    print(df)
    print("\n")
    
    # Diretório para salvar os arquivos
    os.makedirs("/home/giovani/testes", exist_ok=True)
    
    # Salvar o DataFrame em um arquivo Excel
    excel_filename = os.path.join("/home/giovani/testes", f"categorical_discrepancies_{year}.xlsx")
    df.to_excel(excel_filename, index=False)
    
    # Ajustar a largura das colunas e a altura das linhas
    wb = load_workbook(excel_filename)
    ws = wb.active
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                max_height = max(max_height, len(str(cell.value).split('\n')))
        ws.row_dimensions[row[0].row].height = max_height * 15
    
    wb.save(excel_filename)
    print(f"Categorical discrepancies saved to {excel_filename}")


def read_csv_with_encoding(file, encoding):
    """Tenta ler um arquivo CSV com diferentes encodings.
    
    Args:
        file (str): O caminho do arquivo CSV.
        encoding (str): O encoding a ser usado.
        
    Returns:
        DataFrame: O DataFrame lido.
    """
    return pd.read_csv(file, encoding=encoding, delimiter=';', low_memory=False)


def check_compatibility(years):
    """Checa a compatibilidade dos arquivos de entrada com as versões mais recentes.
    
    Args:
        years (list): Lista de anos para os quais os arquivos serão verificados.
    """
    categorical_columns = ['CD_AREA_AVALICAO', 'NM_AREA_AVALICAO', 'CD_CONCEITO_CURSO', 'CD_CONCEITO_PROGRAMA', 'CS_STATUS_JURIDICO', 'DS_DEPENDENCIA_ADMINSTRATIVA', 'DS_FAIXA_ETARIA', 'DS_FAIXA_ETARIA_DISCENTE', 'DS_GRAU_ACADEMICO_DISCENTE', 'DS_TIPO_NACIONALIDADE_DISCENTE', 'NM_GRANDE_AREA_CONHECIMENTO', 'NM_GRAU_PROGRAMA', 'NM_MODALIDADE_PROGRAMA', 'NM_NIVEL_CONCLUSAO_DISCENTE', 'NM_NIVEL_PROGRAMA', 'NM_NIVEL_TITULACAO_DISCENTE', 'NM_REGIAO', 'NM_REGIAO_ENTIDADE', 'NM_SITUACAO_DISCENTE', 'NM_TIPO_DISCENTE_ORIENT_PRINC', 'SG_UF_ENTIDADE_ENSINO', 'SG_UF_PROGRAMA', 'ST_INGRESSANTE', 'TP_DOCUMENTO_DISCENTE']
    
    for year in years:
        files_2021 = read_files(year, '2021-11-10')
        files_2023 = read_files(year, '2023-12-01')
        
        for file_2021, file_2023 in zip(files_2021, files_2023):
            columns_2021 = extract_columns(file_2021)
            columns_2023 = extract_columns(file_2023)
            common_columns, missing_columns, extra_columns, altered_columns = compare_columns(columns_2021, columns_2023)
            print_discrepancies(year, file_2021, file_2023, common_columns, missing_columns, extra_columns, altered_columns)

            encoding = 'latin-1'
            
            df_2021 = read_csv_with_encoding(file_2021, encoding=encoding)
            df_2023 = read_csv_with_encoding(file_2023, encoding=encoding)
            discrepancies = compare_categorical_columns(df_2021, df_2023, categorical_columns)
            print_categorical_discrepancies(year, file_2021, file_2023, discrepancies)

    # Comparar os anos de 2021 e 2022 com a nova versão de 2020
    files_2020 = read_files(2020, '2023-12-01')
    for year in [2021, 2022]:
        files_new = read_files(year, '2023-11-30')
        for file_2020, file_new in zip(files_2020, files_new):
            columns_2020 = extract_columns(file_2020)
            columns_new = extract_columns(file_new)
            common_columns, missing_columns, extra_columns, altered_columns = compare_columns(columns_2020, columns_new)
            print_discrepancies(year, file_2020, file_new, common_columns, missing_columns, extra_columns, altered_columns)
            
            df_2020 = read_csv_with_encoding(file_2020, encoding='latin-1')
            df_new = read_csv_with_encoding(file_new, encoding='latin-1')
            discrepancies = compare_categorical_columns(df_2020, df_new, categorical_columns)
            print_categorical_discrepancies(year, file_2020, file_new, discrepancies)


if __name__ == "__main__":
    check_compatibility(range(2017, 2021))
