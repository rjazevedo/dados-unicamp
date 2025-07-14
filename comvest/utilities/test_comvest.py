import pandas as pd
import os
from pandas import DataFrame
from openpyxl import load_workbook


def read_excel_tables(file_path: str) -> dict:
    """
    Lê todas as planilhas de um arquivo Excel.

    Parâmetros:
    ----------
    file_path : str
        O caminho do arquivo Excel.

    Retorna:
    -------
    dict
        Um dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.
    """
    # Lê todas as planilhas do arquivo Excel
    all_sheets = pd.read_excel(file_path, sheet_name=None)
    
    return all_sheets


def compare_sheets(year: int, sheets_1: dict, sheets_2: dict) -> list[str]:
    """
    Compara os nomes das tabelas de dois arquivos Excel.
    
    Salva as discrepâncias em um novo arquivo Excel.

    Parâmetros:
    ----------
    sheets_1 : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.
    sheets_2 : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.

    Retorna:
    -------
    list
        Uma lista de nomes de planilhas que estão presentes tanto em sheets_1 quanto em sheets_2.
    """
    # Inicializa as listas de planilhas diferentes
    only_in_sheets_1 = []
    only_in_sheets_2 = []
    equal_sheets = []
    
    # Compara as planilhas
    for sheet_name in sheets_1.keys():
        if sheet_name in sheets_2:
            equal_sheets.append(sheet_name)
        else:
            only_in_sheets_1.append(sheet_name)
    
    # Adiciona as planilhas que estão em sheets_2 mas não em sheets_1
    for sheet_name in sheets_2.keys():
        if sheet_name not in sheets_1:
            only_in_sheets_2.append(sheet_name)
    
    # Salvando as discrepâncias em um arquivo Excel
    rows = []
    max_length = max(len(only_in_sheets_1), len(equal_sheets), len(only_in_sheets_2))
    # Pega os dois primeiros dígitos do ano
    prefix = str(year)[:2]
    # Pega os dois últimos dígitos do ano
    sufix = str(year)[2:]
    
    for i in range(max_length):
        row = {
            "Nomes de planilhas iguais": equal_sheets[i] if i < len(equal_sheets) else '',
            f"Somente em 20{prefix}": only_in_sheets_1[i] if i < len(only_in_sheets_1) else '',
            f"Somente em 20{sufix}": only_in_sheets_2[i] if i < len(only_in_sheets_2) else ''
        }
        rows.append(row)
    
    # Criar o DataFrame a partir das linhas
    df = pd.DataFrame(rows)
    
    # Diretório para salvar os arquivos
    os.makedirs("/home/giovani/testes/Comvest", exist_ok=True)
    
    # Salvar o DataFrame em um arquivo Excel
    excel_filename = os.path.join("/home/giovani/testes/Comvest", f"sheets_discrepancies_{year}.xlsx")
    df.to_excel(excel_filename, index=False)
    
    # Ajustar a largura das colunas e a altura das linhas
    wb = load_workbook(excel_filename)
    ws = wb.active
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                max_height = max(max_height, len(str(cell.value).split('\n')))
        ws.row_dimensions[row[0].row].height = max_height * 15
    
    wb.save(excel_filename)
    print(f"Discrepancies saved to {excel_filename}")

    return equal_sheets


def compare_columns(equal_sheets: list[str], sheets_1: dict, sheets_2: dict) -> dict:
    """
    Compara as colunas das planilhas que estão presentes em ambos os arquivos.

    Parâmetros:
    ----------
    equal_sheets : list[str]
        Lista de nomes das planilhas que estão presentes em ambos os arquivos.
    sheets_1 : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.
    sheets_2 : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.

    Retorna:
    -------
    dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são dicionários contendo listas de colunas.
    """
    comparison = {}
    for sheet in equal_sheets:
        columns_1 = set(sheets_1[sheet].columns)
        columns_2 = set(sheets_2[sheet].columns)
        
        missing_columns = columns_1 - columns_2
        extra_columns = columns_2 - columns_1
        common_columns = columns_1 & columns_2
        
        altered_columns = [col for col in common_columns if list(sheets_1[sheet].columns).index(col) != list(sheets_2[sheet].columns).index(col)]
        
        comparison[sheet] = {
            "common_columns": list(common_columns),
            "missing_columns": list(missing_columns),
            "extra_columns": list(extra_columns),
            "altered_columns": altered_columns
        }
        
    return comparison


def print_columns_discrepancies(year: int, equal_sheets: list[str], sheets_1: dict, sheets_2: dict) -> None:
    """
    Salva em arquivos Excel as diferenças entre as colunas existentes para cada uma das chaves no dicionário comparison para o ano indicado.

    Parâmetros:
    ----------
    year : int
        O ano de referência para a comparação.
    equal_sheets : list[str]
        Lista de nomes das planilhas que estão presentes em ambos os arquivos.
    sheets_1 : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.
    sheets_2 : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.
    """
    comparison = compare_columns(equal_sheets, sheets_1, sheets_2)
    
    for sheet_name, columns in comparison.items():
        common_columns = columns["common_columns"]
        missing_columns = columns["missing_columns"]
        extra_columns = columns["extra_columns"]
        altered_columns = columns["altered_columns"]
        
        rows = []
        max_length = max(len(common_columns), len(missing_columns), len(extra_columns), len(altered_columns))
        
        for i in range(max_length):
            row = {
                "Colunas Iguais": common_columns[i] if i < len(common_columns) else '',
                "Colunas Faltantes": missing_columns[i] if i < len(missing_columns) else '',
                "Colunas Extras": extra_columns[i] if i < len(extra_columns) else '',
                "Colunas Alteradas": altered_columns[i] if i < len(altered_columns) else ''
            }
            rows.append(row)
        
        # Criar o DataFrame a partir das linhas
        df = pd.DataFrame(rows)
        
        # Diretório para salvar os arquivos
        output_dir = f"/home/giovani/testes/Comvest/{year}"
        os.makedirs(output_dir, exist_ok=True)
        
        # Salvar o DataFrame em um arquivo Excel
        excel_filename = os.path.join(output_dir, f"columns_discrepancies_{sheet_name}_{year}.xlsx")
        df.to_excel(excel_filename, index=False)
        
        # Ajustar a largura das colunas e a altura das linhas
        wb = load_workbook(excel_filename)
        ws = wb.active
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                if cell.value:
                    max_height = max(max_height, len(str(cell.value).split('\n')))
            ws.row_dimensions[row[0].row].height = max_height * 15
        
        wb.save(excel_filename)
        print(f'Discrepâncias salvas em {excel_filename}')


def organize_comparison_dados(dados_2023: DataFrame, dados_2024: DataFrame, year: int) -> None:
    """
    Ordena as colunas dos dois DataFrames em ordem alfabética.
    O resultado é salvo em um arquivo Excel com as colunas: colunas_2023 e colunas_2024.
    
    Parâmetros
    ----------
    dados_2023 : DataFrame
        DataFrame contendo os dados de 2023.
    dados_2024 : DataFrame
        DataFrame contendo os dados de 2024.
    year : int
        O ano de referência para a comparação.
        
    Retorna
    -------
    None
    """
    rows = []
    # Ordena as colunas dos dois DataFrames em ordem alfabética
    columns_2023 = sorted(dados_2023.columns)
    columns_2024 = sorted(dados_2024.columns)
    # Preenche as linhas com os dados
    max_length = max(len(columns_2023), len(columns_2024))
    for i in range(max_length):
        row = {
            "Colunas 2023": columns_2023[i] if i < len(columns_2023) else '',
            "Colunas 2024": columns_2024[i] if i < len(columns_2024) else ''
        }
        rows.append(row)
    
    # Criar o DataFrame a partir das linhas
    df = pd.DataFrame(rows)
    
    # Diretório para salvar os arquivos
    output_dir = f"/home/giovani/testes/Comvest/{year}"
    os.makedirs(output_dir, exist_ok=True)
    
    # Salvar o DataFrame em um arquivo Excel
    excel_filename = os.path.join(output_dir, f"columns_discrepancies_perfil_{year}.xlsx")
    df.to_excel(excel_filename, index=False)
    
    # Ajustar a largura das colunas e a altura das linhas
    wb = load_workbook(excel_filename)
    ws = wb.active
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                max_height = max(max_height, len(str(cell.value).split('\n')))
        ws.row_dimensions[row[0].row].height = max_height * 15
    
    wb.save(excel_filename)
    print(f'Discrepâncias salvas em {excel_filename}')


def verify_cpfs_Profis(profis_sheets: dict, sheets_2022: dict) -> None:
    """
    Verifica se os CPFs dos candidatos do Profis estão presentes nos arquivos de ingresso do ano indicado.

    Parâmetros:
    ----------
    year : int
        O ano de referência para a comparação.
    profis_sheets : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.
    sheets_2022 : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.
    """
    profis_cpfs = set(profis_sheets["2022"]["cpf"])
    sheets_2022_cpfs = set(sheets_2022["profis_dados"]["CPF"])
    
    missing_cpfs = profis_cpfs - sheets_2022_cpfs
    
    if len(missing_cpfs) > 0:
        print(f"CPFs faltantes em 2022: {missing_cpfs}")
    else:
        print("Todos os CPFs do Profis estão presentes em 2022")


def main() -> None:
    print("Lendo arquivo ingresso2024.xlsx")
    #sheets_24 = read_excel_tables("/home/input/COMVEST/ingresso2024.xlsx")
    """
    print("Lendo arquivo ingresso2021.xlsx")
    # Usado para comparação com os anos seguintes
    sheets_21 = read_excel_tables("/home/input/COMVEST/ingresso2021.xlsx")

    print("Lendo arquivo ingresso2022.xlsx")
    sheets_22 = read_excel_tables("/home/input/COMVEST/ingresso2022.xlsx")

    print("Lendo arquivo ingresso2023.xlsx")
    sheets_23 = read_excel_tables("/home/input/COMVEST/ingresso2023.xlsx")
        
    

    print("Comparando os arquivos de ingresso para os anos de 2021 e 2024")
    equal_sheets_21_24 = compare_sheets(2124, sheets_21, sheets_24)

    print("Comparando os arquivos de ingresso para os anos de 2022 e 2024")
    equal_sheets_22_24 = compare_sheets(2224, sheets_22, sheets_24)

    print("Comparando os arquivos de ingresso para os anos de 2023 e 2024")
    equal_sheets_23_24 = compare_sheets(2324, sheets_23, sheets_24)

    print("Comparando as colunas dos arquivos de ingresso para os anos de 2021 e 2024")
    print_columns_discrepancies(2124, equal_sheets_21_24, sheets_21, sheets_24)

    print("Comparando as colunas dos arquivos de ingresso para os anos de 2022 e 2024")
    print_columns_discrepancies(2224, equal_sheets_22_24, sheets_22, sheets_24)

    print("Comparando as colunas dos arquivos de ingresso para os anos de 2023 e 2024")
    print_columns_discrepancies(2324, equal_sheets_23_24, sheets_23, sheets_24)
    """
    """
    #Lê apenas a planilha "dados" do arquivo de 2023
    dados_2023 = pd.read_excel("/home/input/COMVEST/ingresso2023.xlsx", sheet_name="perfil")
    #Lê apenas a planilha "dados" do arquivo de 2024
    dados_2024 = pd.read_excel("/home/input/COMVEST/ingresso2024.xlsx", sheet_name="perfil")
    organize_comparison_dados(dados_2023, dados_2024, 2324)
    """ 
    # Imprime todas as colunas que contém "mat" no nome na planilhas "notasf2"
    for year in range(2011, 2025):
        print(f"Analisando o ano {year}")
        if year < 2019:
            notasf2 = pd.read_excel(f"/home/input/COMVEST/vest{year}.xlsx", sheet_name="notasf2")
        else:    
            notasf2 = pd.read_excel(f"/home/input/COMVEST/ingresso{year}.xlsx", sheet_name="notasf2")
            
        # Filtra as colunas que contêm "mat" no nome
        filtered_columns = [col for col in notasf2.columns if "mat" in col.lower()]
        print(f"Colunas que contém 'mat' no ano {year}: {filtered_columns}")
        print()
 
if __name__ == "__main__":
    main()
