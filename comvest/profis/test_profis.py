import pandas as pd
import os
from pandas import DataFrame
from openpyxl import load_workbook
import glob

def compare_columns(comvest: dict, profis: dict) -> dict:
    """
    Compara as colunas das planilhas que estão presentes em ambos os arquivos.

    Parâmetros:
    ----------
    equal_sheets : list[str]
        Lista de nomes das planilhas que estão presentes em ambos os arquivos.
    sheets_1 : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.
    sheets_2 : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.

    Retorna:
    -------
    dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são dicionários contendo listas de colunas.
    """
    comparison = {}
    # Transforma todos os nomes das colunas em minúsculas
    comvest = comvest.rename(columns=lambda x: x.lower() if isinstance(x, str) else x)
    profis = profis.rename(columns=lambda x: x.lower() if isinstance(x, str) else x)
    columns_perfil_comvest = set(comvest.columns)
    columns_profis = set(profis.columns)
    
    missing_columns = columns_perfil_comvest - columns_profis
    extra_columns = columns_profis - columns_perfil_comvest
    common_columns = columns_perfil_comvest & columns_profis
    
    altered_columns = [col for col in common_columns if list(columns_perfil_comvest).index(col) != list(columns_profis).index(col)]
    
    comparison["profis_em_comvest"] = {
        "common_columns": list(common_columns),
        "missing_columns": list(missing_columns),
        "extra_columns": list(extra_columns),
        "altered_columns": altered_columns
    }
        
    return comparison


def print_columns_discrepancies(year: int, comvest, profis) -> None:
    """
    Salva em arquivos Excel as diferenças entre as colunas existentes para cada uma das chaves no dicionário comparison para o ano indicado.

    Parâmetros:
    ----------
    year : int
        O ano de referência para a comparação.
    equal_sheets : list[str]
        Lista de nomes das planilhas que estão presentes em ambos os arquivos.
    sheets_1 : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.
    sheets_2 : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.
    """
    comparison = compare_columns(comvest, profis)
    
    for sheet_name, columns in comparison.items():
        common_columns = columns["common_columns"]
        missing_columns = columns["missing_columns"]
        extra_columns = columns["extra_columns"]
        altered_columns = columns["altered_columns"]
        
        rows = []
        max_length = max(len(common_columns), len(missing_columns), len(extra_columns), len(altered_columns))
        
        for i in range(max_length):
            row = {
                "Colunas Iguais": common_columns[i] if i < len(common_columns) else '',
                "Colunas Faltantes": missing_columns[i] if i < len(missing_columns) else '',
                "Colunas Extras": extra_columns[i] if i < len(extra_columns) else '',
                "Colunas Alteradas": altered_columns[i] if i < len(altered_columns) else ''
            }
            rows.append(row)
        
        # Criar o DataFrame a partir das linhas
        df = pd.DataFrame(rows)
        
        # Ordenar as colunas alfabeticamente
        df = df[sorted(df.columns)]
        
        # Diretório para salvar os arquivos
        output_dir = f"/home/giovani/testes/PerfilComvest_Profis/{year}"
        os.makedirs(output_dir, exist_ok=True)
        
        # Salvar o DataFrame em um arquivo Excel
        excel_filename = os.path.join(output_dir, f"columns_discrepancies_{sheet_name}_{year}.xlsx")
        df.to_excel(excel_filename, index=False)
        
        # Ajustar a largura das colunas e a altura das linhas
        wb = load_workbook(excel_filename)
        ws = wb.active
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                if cell.value:
                    max_height = max(max_height, len(str(cell.value).split('\n')))
            ws.row_dimensions[row[0].row].height = max_height * 15
        
        wb.save(excel_filename)
        print(f'Discrepâncias salvas em {excel_filename}')


def organize_comparison_questions(comvest: set[str], profis: set[str], year: int) -> None:
    """
    Organiza e salva as discrepâncias entre as colunas de questões da comvest e do Profis em um arquivo Excel.
    
    Parâmetros
    ----------
    comvest : set[str]
        Conjunto de colunas do DataFrame da comvest.
    profis : set[str]
        Conjunto de colunas do DataFrame do Profis.
    year : int
        O ano de referência para a comparação.
        
    Retorna
    -------
    None
    """
    rows = []
    # Ordena as colunas dos dois DataFrames em ordem alfabética
    questions_comvest = sorted(comvest)
    questions_profis = sorted(profis)
    # Preenche as linhas com os dados
    max_length = max(len(questions_comvest), len(questions_profis))
    for i in range(max_length):
        row = {
            "Questões Comvest": questions_comvest[i] if i < len(questions_comvest) else '',
            "Questões Profis": questions_profis[i] if i < len(questions_profis) else ''
        }
        rows.append(row)
    
    # Criar o DataFrame a partir das linhas
    df = pd.DataFrame(rows)
    
    # Diretório para salvar os arquivos
    output_dir = f"/home/giovani/testes/Profis/{year}"
    os.makedirs(output_dir, exist_ok=True)
    
    # Salvar o DataFrame em um arquivo Excel
    excel_filename = os.path.join(output_dir, f"questions_discrepancies_comvest_profis_{year}.xlsx")
    df.to_excel(excel_filename, index=False)
    
    # Ajustar a largura das colunas e a altura das linhas
    wb = load_workbook(excel_filename)
    ws = wb.active
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                max_height = max(max_height, len(str(cell.value).split('\n')))
        ws.row_dimensions[row[0].row].height = max_height * 15
    
    wb.save(excel_filename)
    print(f'Discrepâncias salvas em {excel_filename}')


def verify_cpfs_Profis(profis_sheets: dict, sheets_2022: dict) -> None:
    """
    Verifica se os CPFs dos candidatos do Profis estão presentes nos arquivos de ingresso do ano indicado.

    Parâmetros:
    ----------
    year : int
        O ano de referência para a comparação.
    profis_sheets : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.
    sheets_2022 : dict
        Dicionário onde as chaves são os nomes das planilhas e os valores são os DataFrames.
    """
    profis_cpfs = set(profis_sheets["2022"]["cpf"])
    sheets_2022_cpfs = set(sheets_2022["profis_dados"]["CPF"])
    
    missing_cpfs = profis_cpfs - sheets_2022_cpfs
    
    if len(missing_cpfs) > 0:
        print(f"CPFs faltantes em 2022: {missing_cpfs}")
    else:
        print("Todos os CPFs do Profis estão presentes em 2022")


def get_question_columns(df: pd.DataFrame) -> set[str]:
    """Extracts column names starting with 'q'."""
    if df is None:
        return set()
    return {col for col in df.columns if isinstance(col, str) and col.lower().startswith('q')}


def match_profis_in_comvest_2022(comvest_file_path: str, output_excel: str) -> None:
    """
    função que procure as pessoas que estão nas planilhas dos profis nas demais planilhas do arquivo de ingresso da comvest para o ano de 2022.
    No caso, as informações do profis que estão na comvest são retiradas da planilha profis_dados (INSC_CAND, NOME_CAND, CPF). 
    Em seguida, preciso saber quantas pessoas há no total nessa parte do profis. 
    Também preciso saber quem está lá e também está nas seguintes planilhas, utilizando as seguintes colunas (entre parênteses) para verificar tal correspondência: 
    ve_dados (INSC, NOME_CAND, CPF), dados (INSC, NOMEOFICM CPF), pefil (insc_cand). 
    Além disso, preciso saber quantas pessoas estão em cada uma dessas planilhas e também quantas pessoas estão em todas elas.
    """
    # Carregar o arquivo Excel
    comvest = pd.read_excel(comvest_file_path, sheet_name=None)
    
    print("Lendo os dados de cada planilha de interesse da comvest")
    # Extrair as planilhas relevantes
    profis_dados = comvest["profis_dados"]
    ve_dados = comvest["ve_dados"]
    dados = comvest["dados"]
    perfil = comvest["perfil"]
    
    print("Número de pessoas na planilha profis_dados")
    print(f"{len(profis_dados)}")
    
    print("Encontrando as pessoas que estão na planilha profis_dados e na planilha ve_dados")
    ve_dados_cpfs = set(ve_dados["CPF"])
    profis_dados_cpfs = set(profis_dados["CPF"])
    profis_dados_cpfs_in_ve_dados = profis_dados_cpfs.intersection(ve_dados_cpfs)
    print(f"Número de pessoas que estão na planilha profis_dados e na planilha ve_dados: {len(profis_dados_cpfs_in_ve_dados)}")
    print("Número de pessoas que estão na planilha profis_dados mas não aparecem na planilha ve_dados")
    print(f"{len(profis_dados_cpfs - ve_dados_cpfs)}")
    
    print("Encontrando as pessoas que estão na planilha profis_dados e na planilha dados")
    dados_cpfs = set(dados["CPF"])
    profis_dados_cpfs_in_dados = profis_dados_cpfs.intersection(dados_cpfs)
    print(f"Número de pessoas que estão na planilha profis_dados e na planilha dados: {len(profis_dados_cpfs_in_dados)}")
    print("Número de pessoas que estão na planilha profis_dados mas não aparecem na planilha dados")
    print(f"{len(profis_dados_cpfs - dados_cpfs)}")
    
    print("Encontrando as pessoas que estão na planilha profis_dados e na planilha perfil")
    perfil_inscs = set(perfil["insc_cand"])
    profis_inscs = set(profis_dados["INSC_CAND"])
    profis_dados_inscs_in_perfil = profis_inscs.intersection(perfil_inscs)
    print(f"Número de pessoas que estão na planilha profis_dados e na planilha perfil: {len(profis_dados_inscs_in_perfil)}")
    print("Número de pessoas que estão na planilha profis_dados mas não aparecem na planilha perfil")
    print(f"{len(profis_inscs - perfil_inscs)}")


def main() -> None:
    # Lê todos os arquivos que estão em /home/giovani/testes/ProfisEnem e os concatena em um único arquivo csv
    # 1) Defina o path da pasta onde estão os CSVs
    pasta = "/home/giovani/testes/ProfisEnem"

    # 2) Use glob para listar todos os arquivos .csv
    padrao = os.path.join(pasta, "*.csv")
    lista_arquivos = glob.glob(padrao)

    # 3) Leia cada arquivo com pandas e guarde em uma lista
    dfs = [pd.read_csv(arquivo) for arquivo in lista_arquivos]

    # 4) Retira todos os DataFrames vazios da lista
    dfs = [df for df in dfs if not df.empty]

    # 5) Concatena todos os DataFrames em um único DataFrame
    df_total = pd.concat(dfs, ignore_index=True)

    df_total.to_csv("/home/giovani/testes/ProfisEnem/total_profis.csv", index=False)    
 
if __name__ == "__main__":
    main()
